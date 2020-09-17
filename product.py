from openpyxl import load_workbook,worksheet
"""this consist all the process of product 
    1: product creation
    2: product insert 
    3: product update
    4: product delete
    this class directly intract with excel file for deleting, updating and inserting"""
class Product:
    #construct for creating object of product with 3 optional parameters
    def __init__(self,name=None,quantity=None,price=None):
        self.name = name
        self.quantity = quantity
        self.price = price
        self.id = None
    
    @staticmethod
    def display_products():
        """listing all the products in excel file"""
        products_file = load_workbook("products.xlsx")
        products = products_file.active
        products_file.close()
        products = list(products)
        print("\n")
        for row in products:
            for cell in row:
                #just for displaying in oredered and same space 
                spaces = 20-len(str(cell.value))
                print(cell.value,end=" "*spaces)
            print("")
        print("\n")
    
    def get_details(self,p_id):
        """get id and find the product, assing all the details to existent object"""
        workbook = load_workbook("products.xlsx")
        products = workbook.active
        #loop for finding specified product
        for row in range(2,products.max_row+1):
            if products[row][0].value == p_id:
                self.id = p_id
                self.row = row
                self.quantity= products[row][2].value
                self.name = products[row][1].value
                self.price = products[row][3].value
                break
        else:
            self.id = 0
            print("no Such Id exits!_ ")
    
    def insert_product(self):
        workbook = load_workbook("products.xlsx")
        worksheet = workbook.worksheets[0]
        r = 0
        #prevent from duplicate entery and check if product with name and  price exist before
        for row in range(2,worksheet.max_row+1):
            if worksheet[row][1].value == self.name and int(worksheet[row][3].value) == int(self.price):
                r = row
                break
        #if product exist befor it wold be updated not inserted
        if r != 0:
            worksheet[row][2].value = int(worksheet[row][2].value)+int(self.quantity)
            print("Updated Successfully")
        else:
            try:
                #getting the last id of products in excel file
                last_id = int(worksheet[worksheet.max_row][0].value)
            except:
                #if it's the first time to write in excel file
                last_id = 0
            last_id +=1
            newrow = [last_id,self.name,self.quantity,self.price]
            worksheet.append(newrow)
            print("insert succesfully!")
        workbook.save("products.xlsx")
        workbook.close()
    #updating the quantity of product
    def update_product(self,quantity):
        workbook = load_workbook("products.xlsx")
        products = workbook.worksheets[0]
        products[self.row][2].value = int(self.quantity) - quantity
        if int(products[self.row][2].value) == 0:
            products.delete_rows(self.row)
        workbook.save("products.xlsx")
        workbook.close()
